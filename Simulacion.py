import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

# ============================================================
# 1. DATOS DEL PROBLEMA
# ============================================================
DIAS = 25
# Costo fijo dado por el problema: 70 dólares diarios
COSTO_FIJO_DOLARES = 70
# Como la producción está expresada en miles de unidades,
# trabajamos los valores monetarios en miles de dólares.
# Por eso:
# 70 dólares = 0.07 miles de dólares.
COSTO_FIJO_MILES = COSTO_FIJO_DOLARES / 1000.0

# ============================================================
# 2. DISTRIBUCIÓN DE PRODUCCIÓN
# ============================================================
producciones = [80, 100, 120]
prob_produccion = [
    0.30,
    0.50,
    0.20
]

# ============================================================
# 3. DISTRIBUCIÓN DE MATERIA PRIMA
# ============================================================
materias = [
    "A",
    "B",
    "C"
]
costos_unitarios = [
    0.55,
    0.65,
    0.75
]
prob_materia = [
    0.40,
    0.35,
    0.25
]

# ============================================================
# 4. DISTRIBUCIÓN DE PRECIO DE VENTA
# ============================================================
precios = [
    1.60,
    1.75,
    2.00
]
prob_precio = [
    0.20,
    0.50,
    0.30
]

# ============================================================
# 5. NÚMEROS ALEATORIOS DEL EJERCICIO
# ============================================================
# ------------------------------------------------------------
# Producción
# RN de un dígito
# ------------------------------------------------------------
rn_produccion = [
    3, 8, 1, 9, 0,
    6, 2, 5, 3, 7,
    0, 4, 8, 1, 9,
    2, 6, 7, 5, 0,
    4, 8, 3, 2, 8
]
# ------------------------------------------------------------
# Materia prima
# RN entre 1 y 100
# ------------------------------------------------------------
rn_materia = [
    83, 17, 45, 92, 38,
    74, 59, 21, 66, 10,
    87, 33, 61, 48, 95,
    12, 70, 27, 99, 36,
    54, 76, 68, 19, 41
]
# ------------------------------------------------------------
# Precio de venta
# RN de un dígito
# ------------------------------------------------------------
rn_precio = [
    4, 2, 7, 1, 0,
    3, 5, 6, 9, 8,
    4, 1, 7, 3, 0,
    2, 8, 9, 6, 5,
    7, 2, 0, 3, 1
]

# ============================================================
# 6. CALCULAR PROBABILIDADES ACUMULADAS
# ============================================================
def calcular_acumuladas(probabilidades):
    acumuladas = []
    acumulada = 0
    for p in probabilidades:
        acumulada = acumulada + p
        acumuladas.append(round(acumulada, 10))
    return acumuladas

acum_produccion = calcular_acumuladas(prob_produccion)
acum_materia = calcular_acumuladas(prob_materia)
acum_precio = calcular_acumuladas(prob_precio)

# ============================================================
# 7. CONSTRUIR INTERVALOS PARA RN DE UN DÍGITO
# ============================================================
def construir_intervalos_digito(probabilidades):
    """
    Se trabaja con el universo:
    1,2,3,4,5,6,7,8,9,0
    Ejemplo:
    probabilidad 0.30
    => 3 posiciones
    => 1,2,3
    """
    universo = [
        1, 2, 3, 4, 5,
        6, 7, 8, 9, 0
    ]
    intervalos = []
    posicion = 0

    for p in probabilidades:
        cantidad = round(p * 10)
        numeros = universo[posicion:posicion + cantidad]
        intervalos.append(numeros)
        posicion = posicion + cantidad
    return intervalos
# Producción
intervalos_produccion = construir_intervalos_digito(
    prob_produccion)
# Precio
intervalos_precio = construir_intervalos_digito(
    prob_precio)

# ============================================================
# 8. CONSTRUIR INTERVALOS PARA RN DE 1 A 100
# ============================================================
def construir_intervalos_100(probabilidades):
    intervalos = []
    inicio = 1
    for p in probabilidades:
        cantidad = round(p * 100)
        fin = inicio + cantidad - 1
        intervalos.append((inicio, fin))
        inicio = fin + 1
    return intervalos

intervalos_materia = construir_intervalos_100(
    prob_materia)

# ============================================================
# 9. MOSTRAR INTERVALOS COMO TEXTO
# ============================================================
def texto_intervalo_digito(numeros):
    """
    Ejemplos:
    [1,2,3] -> 1-3
    [4,5,6,7,8] -> 4-8
    [9,0] -> 9-0
    """
    if len(numeros) == 1:
        return str(numeros[0])
    return f"{numeros[0]}-{numeros[-1]}"

def texto_intervalo_100(intervalo):
    inicio, fin = intervalo
    return f"{inicio}-{fin}"

# ============================================================
# 10. TRANSFORMAR RN EN VALORES
# ============================================================
def obtener_valor_digito(rn, valores, intervalos):
    for valor, conjunto in zip(valores, intervalos):
        if rn in conjunto:
            return valor
    raise ValueError(
        f"El RN {rn} no pertenece "
        "a ningún intervalo.")

def obtener_valor_100(rn, valores, intervalos):
    for valor, intervalo in zip(valores, intervalos):
        inicio, fin = intervalo
        if inicio <= rn <= fin:
            return valor
    raise ValueError(
        f"El RN {rn} no pertenece "
        "a ningún intervalo.")

# ============================================================
# 11. SIMULACIÓN DE LOS 25 DÍAS
# ============================================================
resultados = []
for i in range(DIAS):
    dia = i + 1
    # --------------------------------------------------------
    # PRODUCCIÓN
    # --------------------------------------------------------
    rn_prod = rn_produccion[i]
    produccion = obtener_valor_digito(rn_prod, producciones,
        intervalos_produccion)
    # --------------------------------------------------------
    # PRECIO
    # --------------------------------------------------------
    rn_prec = rn_precio[i]
    precio = obtener_valor_digito(rn_prec, precios,
        intervalos_precio)
    # --------------------------------------------------------
    # INGRESO
    # --------------------------------------------------------
    ingreso = produccion * precio
    # --------------------------------------------------------
    # MATERIA PRIMA
    # --------------------------------------------------------
    rn_mat = rn_materia[i]
    materia = obtener_valor_100(rn_mat, materias,
        intervalos_materia)
    # --------------------------------------------------------
    # COSTO UNITARIO
    # --------------------------------------------------------
    indice_materia = materias.index(materia)
    costo_unitario = costos_unitarios[indice_materia]
    # --------------------------------------------------------
    # COSTO VARIABLE
    # --------------------------------------------------------
    costo_variable = (produccion * costo_unitario)
    # --------------------------------------------------------
    # COSTO TOTAL
    # --------------------------------------------------------
    costo_total = (costo_variable + COSTO_FIJO_MILES)
    # --------------------------------------------------------
    # GANANCIA
    # --------------------------------------------------------
    ganancia = (ingreso - costo_total)
    # --------------------------------------------------------
    # GUARDAR RESULTADO DEL DÍA
    # --------------------------------------------------------
    resultados.append({
        "Dia": dia,
        "RN Produccion": rn_prod,
        "Produccion": produccion,
        "RN Precio": rn_prec,
        "Precio Venta": precio,
        "Ingreso": ingreso,
        "RN Materia": rn_mat,
        "Materia Prima": materia,
        "Costo Unitario": costo_unitario,
        "Costo Variable": costo_variable,
        "Costo Total": costo_total,
        "Ganancia": ganancia
    })

# ============================================================
# 12. RESULTADOS GENERALES
# ============================================================
ingreso_total = sum(fila["Ingreso"] for fila in resultados)
costo_total_general = sum(fila["Costo Total"] 
    for fila in resultados)
ganancia_total = sum(fila["Ganancia"]
    for fila in resultados)
produccion_promedio = sum(fila["Produccion"]
    for fila in resultados) / DIAS
precio_promedio = sum(fila["Precio Venta"]
    for fila in resultados) / DIAS
ganancia_promedio = (ganancia_total / DIAS)

# ============================================================
# 13. CREAR EL EXCEL
# ============================================================
wb = Workbook()
# ============================================================
# HOJA 1: SIMULACIÓN
# ============================================================
ws = wb.active
ws.title = "Simulacion"
# -----------------------------------------------------------
# TÍTULO
# ------------------------------------------------------------
ws.merge_cells("A1:L1")
ws["A1"] = ("SIMULACIÓN DE PRODUCCIÓN Y VENTAS")
ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
ws["A1"].alignment = Alignment(horizontal="center")
ws.merge_cells("A2:L2")
ws["A2"] = ("Simulación de 25 días de operación")
ws["A2"].font = Font(italic=True, color="555555")
ws["A2"].alignment = Alignment(horizontal="center")
# ------------------------------------------------------------
# ENCABEZADOS
# ------------------------------------------------------------
encabezados = [
    "Día",
    "RN Producción",
    "Producción",
    "RN Precio",
    "Precio Venta",
    "Ingreso",
    "RN Materia",
    "Materia Prima",
    "Costo Unitario",
    "Costo Variable",
    "Costo Total",
    "Ganancia"
]
fila_encabezado = 4
for columna, encabezado in enumerate(
    encabezados, start=1):
    celda = ws.cell(
        row=fila_encabezado,
        column=columna,
        value=encabezado
        )
    celda.font = Font(bold=True, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor="4472C4")
    celda.alignment = Alignment(
        horizontal="center", vertical="center"
        )
# ------------------------------------------------------------
# DATOS DE LOS 25 DÍAS
# ------------------------------------------------------------
for fila_excel, resultado in enumerate(resultados, start=5):
    ws.cell(fila_excel, 1, resultado["Dia"])
    ws.cell(fila_excel, 2, resultado["RN Produccion"])
    ws.cell(fila_excel, 3, resultado["Produccion"])
    ws.cell(fila_excel, 4, resultado["RN Precio"])
    ws.cell(fila_excel, 5, resultado["Precio Venta"])
    ws.cell(fila_excel, 6, resultado["Ingreso"])
    ws.cell(fila_excel, 7, resultado["RN Materia"])
    ws.cell(fila_excel, 8, resultado["Materia Prima"])
    ws.cell(fila_excel, 9, resultado["Costo Unitario"])
    ws.cell(fila_excel, 10, resultado["Costo Variable"])
    ws.cell(fila_excel, 11, resultado["Costo Total"])
    ws.cell(fila_excel, 12, resultado["Ganancia"])
# ------------------------------------------------------------
# FORMATO NUMÉRICO
# ------------------------------------------------------------
for fila in range(5, 5 + DIAS):
    ws.cell(fila, 5).number_format = "0.00"
    ws.cell(fila, 6).number_format = "0.00"
    ws.cell(fila, 9).number_format = "0.00"
    ws.cell(fila, 10).number_format = "0.00"
    ws.cell(fila, 11).number_format = "0.00"
    ws.cell(fila, 12).number_format = "0.00"
# ============================================================
# HOJA 2: DISTRIBUCIONES
# ============================================================
wd = wb.create_sheet("Distribuciones")
# ------------------------------------------------------------
# PRODUCCIÓN
# ------------------------------------------------------------
wd.merge_cells("A1:D1")
wd["A1"] = ("DISTRIBUCIÓN DE PRODUCCIÓN")
wd["A1"].font = Font(bold=True, size=14, color="FFFFFF")
wd["A1"].fill = PatternFill("solid", fgColor="70AD47")
wd["A1"].alignment = Alignment(horizontal="center")
enc_prod = [
    "Producción",
    "Probabilidad",
    "Acumulada",
    "Intervalo RN"
]
for col, texto in enumerate(enc_prod, 1):
    wd.cell(2, col, texto)
for i in range(len(producciones)):
    wd.cell(3 + i, 1, producciones[i])
    wd.cell(3 + i, 2, prob_produccion[i])
    wd.cell(3 + i, 3, acum_produccion[i])
    wd.cell(3 + i, 4,
        texto_intervalo_digito(intervalos_produccion[i]))
# ------------------------------------------------------------
# MATERIA PRIMA
# ------------------------------------------------------------
wd.merge_cells("A8:E8")
wd["A8"] = ("DISTRIBUCIÓN DE MATERIA PRIMA")
wd["A8"].font = Font(bold=True, size=14, color="FFFFFF")
wd["A8"].fill = PatternFill("solid", fgColor="ED7D31")
wd["A8"].alignment = Alignment(horizontal="center")
enc_mat = [
    "Materia",
    "Costo Unitario",
    "Probabilidad",
    "Acumulada",
    "Intervalo RN"
]
for col, texto in enumerate(enc_mat, 1):
    wd.cell( 9, col, texto) 
for i in range(len(materias)):
    wd.cell(10 + i, 1, materias[i])
    wd.cell(10 + i, 2, costos_unitarios[i])
    wd.cell(10 + i, 3, prob_materia[i])
    wd.cell(10 + i, 4, acum_materia[i])
    wd.cell(10 + i, 5,
        texto_intervalo_100(intervalos_materia[i]))
# ------------------------------------------------------------
# PRECIO DE VENTA
# ------------------------------------------------------------
wd.merge_cells("A15:D15")
wd["A15"] = ("DISTRIBUCIÓN DEL PRECIO DE VENTA")
wd["A15"].font = Font(bold=True, size=14, color="FFFFFF")
wd["A15"].fill = PatternFill( "solid", fgColor="5B9BD5")
wd["A15"].alignment = Alignment(horizontal="center")
enc_precio = ["Precio", "Probabilidad", "Acumulada", "Intervalo RN"]
for col, texto in enumerate(enc_precio, 1):
    wd.cell(16, col, texto)
for i in range(len(precios)):
    wd.cell(17 + i, 1, precios[i])
    wd.cell(17 + i, 2, prob_precio[i])
    wd.cell(17 + i, 3, acum_precio[i])
    wd.cell(17 + i, 4,
        texto_intervalo_digito(intervalos_precio[i]))
# ============================================================
# HOJA 3: RESUMEN
# ============================================================
wr = wb.create_sheet("Resumen")
wr.merge_cells("A1:B1")
wr["A1"] = ("RESUMEN DE LA SIMULACIÓN")
wr["A1"].font = Font(bold=True, size=16, color="FFFFFF")
wr["A1"].fill = PatternFill("solid", fgColor="7030A0")
wr["A1"].alignment = Alignment(horizontal="center")
resumen = [
    ("Número de días", DIAS),
    ("Costo fijo diario ($)", COSTO_FIJO_DOLARES),
    ("Ingreso total", ingreso_total),
    ("Costo total", costo_total_general),
    ("Ganancia total", ganancia_total),
    ("Ganancia promedio diaria", ganancia_promedio),
    ("Producción promedio", produccion_promedio),
    ("Precio promedio", precio_promedio)
]
for fila, (nombre, valor) in enumerate(resumen, start=3):
    wr.cell(fila, 1, nombre)
    wr.cell(fila, 2, valor)
    wr.cell(fila, 1).font = Font(bold=True)
for fila in range(5, 11):
    wr.cell(fila, 2).number_format = "0.00"

# ============================================================
# 14. GRÁFICO DE GANANCIA DIARIA
# ============================================================
grafico = BarChart()
grafico.title = ("Ganancia diaria")
grafico.y_axis.title = ("Ganancia")
grafico.x_axis.title = ("Día")
datos_grafico = Reference(ws, min_col=12, min_row=4, max_row=4 + DIAS)
categorias = Reference(ws, min_col=1, min_row=5, max_row=4 + DIAS)
grafico.add_data(datos_grafico, titles_from_data=True)
grafico.set_categories(categorias)
grafico.height = 8
grafico.width = 16
wr.add_chart(grafico, "D3")

# ============================================================
# 15. BORDES Y FORMATO GENERAL
# ============================================================
borde = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9")
    )
for hoja in [ws, wd, wr]:
    for fila in hoja.iter_rows():
        for celda in fila:
            if celda.value is not None:
                celda.border = borde
                celda.alignment = Alignment(
                    vertical="center")

# ============================================================
# 16. AJUSTAR ANCHO DE COLUMNAS
# ============================================================
for hoja in [ws,wd, wr]:
    for columna in hoja.columns:
        max_longitud = 0
        letra = get_column_letter(
            columna[0].column)
        for celda in columna:
            if celda.value is not None:
                longitud = len(str(celda.value))
                if longitud > max_longitud:
                    max_longitud = longitud
        hoja.column_dimensions[letra].width = min(
            max_longitud + 3, 25)
# Congelar encabezados
ws.freeze_panes = "A5"

# ============================================================
# 17. GUARDAR EL EXCEL EN LA MISMA CARPETA DEL .PY
# ============================================================
# __file__ representa la ubicación de este archivo Python.
carpeta_script = os.path.dirname(
    os.path.abspath(__file__)
)
# Construimos la ruta completa del Excel.
nombre_archivo = os.path.join(
    carpeta_script,
    "Simulacion_Produccion_Ventas.xlsx")
# Guardamos el libro.
wb.save(nombre_archivo)

# ============================================================
# 18. RESULTADOS EN CONSOLA
# ============================================================
print("=" * 60)
print("SIMULACIÓN COMPLETADA")
print("=" * 60)
print(f"Días procesados: {DIAS}")
print(
    f"Ingreso total: "
    f"{ingreso_total:.2f}"
)
print(
    f"Costo total: "
    f"{costo_total_general:.2f}"
)
print(
    f"Ganancia total: "
    f"{ganancia_total:.2f}"
)
print(
    f"Ganancia promedio: "
    f"{ganancia_promedio:.2f}"
)
print(
    f"Producción promedio: "
    f"{produccion_promedio:.2f}"
)
print(
    f"Precio promedio: "
    f"{precio_promedio:.3f}"
)
print()
print("Excel guardado en:")
print(nombre_archivo)
print()
print("¿El archivo existe?:", os.path.exists(nombre_archivo))

# ============================================================
# 19. ABRIR AUTOMÁTICAMENTE EL EXCEL
# ============================================================
if os.path.exists(nombre_archivo):
    os.startfile(nombre_archivo)